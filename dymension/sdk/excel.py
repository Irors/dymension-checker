import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

class Excel:

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Address'
    sheet['B1'] = 'Number of transactions'

    column = sheet.column_dimensions[openpyxl.utils.get_column_letter(1)]
    column.width = 16

    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(2)]
    column_B.width = 20

    alignment_style = Alignment(horizontal='center', vertical='center', indent=3, wrap_text=True)

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for i in ('AB'):
        cell = sheet[f'{i}1']
        bold_font = Font(bold=True)
        cell.font = bold_font

        cell.alignment = alignment_style
        cell.border = border_style
